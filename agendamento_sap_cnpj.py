import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import json
import threading
from datetime import datetime
from pathlib import Path

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    import time
    WATCHDOG_OK = True
except ImportError:
    WATCHDOG_OK = False

CONFIG_PATH = Path(__file__).parent / "agendamento_sap_cnpj_config.json"

# ─────────────────────────────────────────────────────────────────────────────
#  PROCESSADOR
# ─────────────────────────────────────────────────────────────────────────────
class ProcessadorAgendamento:

    def ler(self, path, ctx):
        """Le Excel ou TXT/CSV com separador pipe (SAP) ou virgula/ponto-e-virgula."""
        ext = Path(path).suffix.lower()
        if ext in ('.xlsx', '.xls', '.xlsb', '.xlsm'):
            try:
                return pd.read_excel(path, dtype=str)
            except Exception:
                sheets = pd.read_excel(path, sheet_name=None, dtype=str)
                return list(sheets.values())[0]

        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        import io
        for enc in encodings:
            try:
                with open(path, 'r', encoding=enc) as f:
                    lines = f.readlines()
                if any('|' in l for l in lines[:5]):
                    clean = [l for l in lines if not all(c in '-+| \t\n' for c in l)]
                    df = pd.read_csv(io.StringIO(''.join(clean)), sep='|',
                                     engine='python', on_bad_lines='skip', dtype=str)
                    df = df.dropna(axis=1, how='all')
                    df.columns = [str(c).strip() for c in df.columns]
                    df = df.apply(lambda s: s.str.strip() if s.dtype == 'object' else s)
                    return df
                return pd.read_csv(path, sep=None, engine='python',
                                   on_bad_lines='skip', encoding=enc, dtype=str)
            except Exception:
                continue
        raise ValueError(f"Nao foi possivel ler '{ctx}'. Verifique o formato.")

    def col(self, df, opcoes, ctx, obrigatorio=True):
        """Localiza coluna por lista de possiveis nomes (case-insensitive, sem espacos extras)."""
        mapa = {str(c).strip().lower(): c for c in df.columns}
        for o in opcoes:
            if o.lower() in mapa:
                return mapa[o.lower()]
        if obrigatorio:
            raise ValueError(
                f"Coluna nao encontrada em '{ctx}'.\n"
                f"Procurei: {opcoes}\n"
                f"Disponiveis: {list(df.columns)}"
            )
        return None

    def processar(self, arqs: dict, pasta_out: str, prefixo: str):
        """
        arqs = {
            'sap':  caminho SAP Export (.xlsx/.txt),
            'cnpj': caminho tabela centro -> CNPJ,
        }
        Retorna (caminho_saida, n_linhas)
        """

        # -- 1. SAP Export ──────────────────────────────────────────────────
        sap = self.ler(arqs['sap'], 'SAP Export')

        c_ped      = self.col(sap, ['Pedido de compras', 'Pedido', 'PO', 'Doc.compra'], 'SAP Export')
        c_forn     = self.col(sap, ['Descricao do fornecedor', 'Descrição do fornecedor',
                                    'Nome emissor', 'Fornecedor', 'Vendor Name',
                                    'Fornecedor/centro fornecedor'], 'SAP Export')
        c_mat      = self.col(sap, ['Codigo do material', 'Código do material',
                                    'Material', 'Cod. Material', 'Cód. Material',
                                    'Cod Material'], 'SAP Export')
        c_desc     = self.col(sap, ['Descricao do material', 'Descrição do material',
                                    'Desc. Material', 'Desc Material', 'Texto breve'], 'SAP Export')
        c_data     = self.col(sap, ['Data doc.', 'Data Documento', 'Data Doc.', 'Posting Date'], 'SAP Export')
        c_centro   = self.col(sap, ['Entregue centro', 'Centro', 'Plant', 'Entregue em', 'Cen.'], 'SAP Export')
        c_contrato = self.col(sap, ['Contrato', 'Contrato de compra',
                                    'Scheduling Agreement', 'SA'], 'SAP Export', obrigatorio=False)
        c_ean      = self.col(sap, ['EAN', 'EAN/UPC', 'Codigo EAN', 'Código EAN',
                                    'GTIN'], 'SAP Export', obrigatorio=False)

        cols_sap = [c for c in [c_ped, c_forn, c_mat, c_desc, c_data, c_centro, c_contrato, c_ean] if c]
        sap_df = sap[cols_sap].copy()

        # -- 2. Tabela CNPJ por Centro ──────────────────────────────────────
        cnpj_tab = self.ler(arqs['cnpj'], 'CNPJ Centros')

        c_cent_tab  = self.col(cnpj_tab, ['Centro', 'Cod Centro', 'Cód Centro',
                                           'Plant', 'Codigo Centro', 'Código Centro'], 'CNPJ Centros')
        c_cnpj_tab  = self.col(cnpj_tab, ['CNPJ', 'CNPJ Destinatario', 'CNPJ Destinatário',
                                           'CNPJ Dest', 'CNPJ do Centro'], 'CNPJ Centros')
        c_cnpj_emit = self.col(cnpj_tab, ['CNPJ Emissor', 'CNPJ Emit', 'Emissor'],
                               'CNPJ Centros', obrigatorio=False)

        cnpj_df = cnpj_tab[[c_cent_tab, c_cnpj_tab] + ([c_cnpj_emit] if c_cnpj_emit else [])].copy()
        cnpj_df['__CENTRO'] = cnpj_df[c_cent_tab].astype(str).str.strip()

        # -- 3. Cruzamento SAP + CNPJ por Centro ───────────────────────────
        sap_df['__CENTRO'] = sap_df[c_centro].astype(str).str.strip()
        merged = pd.merge(sap_df, cnpj_df, on='__CENTRO', how='left')

        # -- 4. Montar arquivo final ───────────────────────────────────────
        out = pd.DataFrame()
        out['Codigo Pedido']        = merged[c_ped]
        out['Contrato']             = merged[c_contrato] if c_contrato else ''
        out['Fornecedor']           = merged[c_forn]
        out['Centro']               = merged[c_centro]
        out['CNPJ Emissor']         = merged[c_cnpj_emit] if c_cnpj_emit else merged[c_cnpj_tab]
        out['CNPJ Destinatario']    = merged[c_cnpj_tab]
        out['Data da Emissao']      = merged[c_data]
        out['EAN']                  = merged[c_ean] if c_ean else ''
        out['Codigo do Produto']    = merged[c_mat]
        out['Descricao do Produto'] = merged[c_desc]

        out = out.drop_duplicates()

        # -- 5. Salvar ─────────────────────────────────────────────────────
        os.makedirs(pasta_out, exist_ok=True)
        ts = datetime.now().strftime('%d%m%Y_%H%M%S')
        dest = os.path.join(pasta_out, f"{prefixo}_{ts}.xlsx")

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agendamento"

        header_fill = PatternFill('solid', start_color='003366', end_color='003366')
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(out.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

        data_font = Font(name='Arial', size=9)
        alt_fill  = PatternFill('solid', start_color='EEF2F7', end_color='EEF2F7')

        for row_idx, row in enumerate(out.itertuples(index=False), 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill

        col_widths = {
            'Codigo Pedido': 16, 'Contrato': 14, 'Fornecedor': 30,
            'Centro': 12, 'CNPJ Emissor': 22, 'CNPJ Destinatario': 22,
            'Data da Emissao': 16, 'EAN': 16, 'Codigo do Produto': 18,
            'Descricao do Produto': 35,
        }
        for col_idx, col_name in enumerate(out.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        wb.save(dest)
        return dest, len(out)


# ─────────────────────────────────────────────────────────────────────────────
#  WATCHER
# ─────────────────────────────────────────────────────────────────────────────
class Watcher(FileSystemEventHandler):
    def __init__(self, monitorados, callback):
        super().__init__()
        self._alvos = {str(Path(p).resolve()) for p in monitorados if p}
        self._cb = callback
        self._last = 0

    def on_modified(self, event):
        if event.is_directory:
            return
        if str(Path(event.src_path).resolve()) in self._alvos:
            agora = time.time()
            if agora - self._last > 6:
                self._last = agora
                self._cb(Path(event.src_path).name)


# ─────────────────────────────────────────────────────────────────────────────
#  INTERFACE
# ─────────────────────────────────────────────────────────────────────────────
class App:
    COR_HEADER  = '#003366'
    COR_ACCENT  = '#1a73e8'
    COR_BG      = '#f0f2f5'
    COR_SUCCESS = '#2e7d32'
    COR_WARN    = '#c62828'

    LABELS = {
        'sap':  ('SAP Export',             True,  '*.xlsx *.xls *.xlsb *.txt *.csv'),
        'cnpj': ('Tabela CNPJ por Centro', True,  '*.xlsx *.xls *.xlsb *.csv'),
    }

    def __init__(self, root):
        self.root = root
        self.root.title("Portal de Agendamento — SAP x CNPJ por Centro")
        self.root.geometry("720x500")
        self.root.configure(bg=self.COR_BG)
        self.root.resizable(False, False)
        self.proc = ProcessadorAgendamento()
        self.cfg  = self._load_cfg()
        self._obs = None
        self._busy = False
        self._build()
        self._start_watcher()

    # -- Config ────────────────────────────────────────────────────────────
    def _load_cfg(self):
        if CONFIG_PATH.exists():
            with open(CONFIG_PATH, encoding='utf-8') as f:
                return json.load(f)
        return {'arquivos': {'sap': '', 'cnpj': ''},
                'output': {'pasta': '', 'prefixo': 'Agendamento_SAP_CNPJ'}}

    def _save_cfg(self):
        for k, v in self.entries.items():
            self.cfg['arquivos'][k] = v.get().strip()
        self.cfg['output']['pasta'] = self.var_pasta.get().strip()
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(self.cfg, f, ensure_ascii=False, indent=2)

    # -- UI ────────────────────────────────────────────────────────────────
    def _build(self):
        hdr = tk.Frame(self.root, bg=self.COR_HEADER, height=64)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)
        tk.Label(hdr, text='Agendamento — SAP Export x CNPJ por Centro',
                 fg='white', bg=self.COR_HEADER,
                 font=('Segoe UI', 14, 'bold')).place(relx=.5, rely=.5, anchor='center')

        body = tk.Frame(self.root, bg=self.COR_BG, padx=32, pady=16)
        body.pack(fill='both', expand=True)

        self.frm_status = tk.Frame(body, bd=1, relief='solid')
        self.frm_status.pack(fill='x', pady=(0, 14))
        self.lbl_status = tk.Label(self.frm_status, font=('Segoe UI', 9), pady=6, padx=10, anchor='w')
        self.lbl_status.pack(fill='x')
        self._set_status(False)

        tk.Label(body, text='Bases de dados', bg=self.COR_BG,
                 font=('Segoe UI', 10, 'bold'), fg='#333').pack(anchor='w')

        self.entries = {}
        for key, (label, req, ext) in self.LABELS.items():
            row = tk.Frame(body, bg=self.COR_BG, pady=5)
            row.pack(fill='x')
            txt = label + (' *' if req else '')
            tk.Label(row, text=txt, bg=self.COR_BG, font=('Segoe UI', 9),
                     width=30, anchor='w').pack(side='left')
            var = tk.StringVar(value=self.cfg['arquivos'].get(key, ''))
            tk.Entry(row, textvariable=var, font=('Segoe UI', 9), width=28
                     ).pack(side='left', padx=(0, 5))
            tk.Button(row, text='...', width=3, relief='flat', bg='#dce3ec',
                      cursor='hand2',
                      command=lambda k=key, v=var, e=ext: self._browse(k, v, e)
                      ).pack(side='left')
            self.entries[key] = var

        tk.Frame(body, bg='#ccd3dc', height=1).pack(fill='x', pady=12)

        row_o = tk.Frame(body, bg=self.COR_BG)
        row_o.pack(fill='x')
        tk.Label(row_o, text='Pasta de saida *', bg=self.COR_BG,
                 font=('Segoe UI', 9), width=30, anchor='w').pack(side='left')
        self.var_pasta = tk.StringVar(value=self.cfg['output'].get('pasta', ''))
        tk.Entry(row_o, textvariable=self.var_pasta, font=('Segoe UI', 9),
                 width=28).pack(side='left', padx=(0, 5))
        tk.Button(row_o, text='...', width=3, relief='flat', bg='#dce3ec',
                  cursor='hand2', command=self._browse_pasta).pack(side='left')

        btn_row = tk.Frame(body, bg=self.COR_BG, pady=14)
        btn_row.pack(fill='x')
        tk.Button(btn_row, text='Salvar configuracao',
                  command=self._salvar_e_reiniciar,
                  bg='#546e7a', fg='white', relief='flat',
                  font=('Segoe UI', 9, 'bold'), padx=14, pady=8,
                  cursor='hand2').pack(side='left')

        self.btn = tk.Button(btn_row, text='  Gerar agora',
                             command=self._gerar,
                             bg=self.COR_ACCENT, fg='white', relief='flat',
                             font=('Segoe UI', 10, 'bold'), padx=18, pady=8,
                             cursor='hand2')
        self.btn.pack(side='right')

        self.log = tk.Text(body, height=6, font=('Courier New', 8),
                           bg='#1e1e2e', fg='#a6e3a1',
                           state='disabled', relief='flat')
        self.log.pack(fill='x')

    def _browse(self, key, var, ext):
        p = filedialog.askopenfilename(filetypes=[('Arquivos', ext)])
        if p:
            var.set(p)

    def _browse_pasta(self):
        p = filedialog.askdirectory()
        if p:
            self.var_pasta.set(p)

    # -- Watcher ───────────────────────────────────────────────────────────
    def _start_watcher(self):
        if not WATCHDOG_OK:
            self._log('watchdog nao instalado - monitoramento automatico desativado.')
            return
        arqs = list(self.cfg['arquivos'].values())
        pastas = {str(Path(p).parent) for p in arqs if p and Path(p).exists()}
        if not pastas:
            self._log('Configure os arquivos e salve para ativar o monitoramento.')
            return
        handler = Watcher(arqs, lambda nome: (
            self._log(f'Alteracao detectada: {nome} - processando...'),
            self.root.after(500, self._gerar)
        ))
        self._obs = Observer()
        for p in pastas:
            self._obs.schedule(handler, p, recursive=False)
        self._obs.start()
        self._set_status(True)
        self._log(f'Monitorando {len(arqs)} arquivo(s).')

    def _stop_watcher(self):
        if self._obs and self._obs.is_alive():
            self._obs.stop()
            self._obs.join()
            self._obs = None

    def _set_status(self, ativo):
        if ativo:
            self.lbl_status.config(
                text='Monitoramento ativo - relatorio gerado automaticamente ao detectar alteracoes',
                fg=self.COR_SUCCESS, bg='#e8f5e9')
            self.frm_status.config(bg='#e8f5e9')
        else:
            self.lbl_status.config(
                text='Monitoramento inativo', fg=self.COR_WARN, bg='#ffebee')
            self.frm_status.config(bg='#ffebee')

    # -- Processamento ─────────────────────────────────────────────────────
    def _salvar_e_reiniciar(self):
        self._save_cfg()
        self._stop_watcher()
        self._start_watcher()
        self._log('Configuracao salva. Monitoramento reiniciado.')

    def _gerar(self):
        if self._busy:
            self._log('Ja em execucao, aguarde...')
            return
        self._busy = True
        self.btn.config(state='disabled', text='Processando...')

        def run():
            try:
                arqs  = dict(self.cfg['arquivos'])
                pasta = self.cfg['output']['pasta']
                pref  = self.cfg['output'].get('prefixo', 'Agendamento_SAP_CNPJ')
                falta = [k for k in ('sap', 'cnpj') if not arqs.get(k)]
                if falta or not pasta:
                    raise ValueError(f"Configure e salve antes de gerar.\nFaltando: {falta or ['pasta de saida']}")
                dest, n = self.proc.processar(arqs, pasta, pref)
                self._log(f'Gerado: {Path(dest).name} ({n} linhas)')
                self.root.after(0, lambda: messagebox.showinfo(
                    'Concluido', f'Arquivo gerado com sucesso!\n\n{dest}\n\n{n} linhas exportadas.'))
            except Exception as e:
                self._log(f'Erro: {e}')
                self.root.after(0, lambda: messagebox.showerror('Erro', str(e)))
            finally:
                self._busy = False
                self.root.after(0, lambda: self.btn.config(
                    state='normal', text='  Gerar agora'))

        threading.Thread(target=run, daemon=True).start()

    def _log(self, msg):
        def _w():
            self.log.config(state='normal')
            ts = datetime.now().strftime('%H:%M:%S')
            self.log.insert('end', f'[{ts}] {msg}\n')
            self.log.see('end')
            self.log.config(state='disabled')
        self.root.after(0, _w)

    def on_close(self):
        self._stop_watcher()
        self.root.destroy()


if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.protocol('WM_DELETE_WINDOW', app.on_close)
    root.mainloop()
